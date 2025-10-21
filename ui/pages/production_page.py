# app/ui/pages/production_page.py
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, date
from ui.components.charts import ChartComponents
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ProductionPage:
    """生産計画ページ（シミュレーション + CRUD管理）"""

    def __init__(self, production_service, transport_service=None):
        self.service = production_service
        self.transport_service = transport_service
        self.charts = ChartComponents()

    # -----------------------------
    # Entry
    # -----------------------------
    def show(self):
        st.title("🏭 生産計画")

        tab1, tab2, tab3 = st.tabs(["📊 計画シミュレーション", "📝 生産計画管理", "🔧 製造工程（加工対象）"])

        with tab1:
            self._show_plan_simulation()

        with tab2:
            self._show_plan_management()

        with tab3:
            self._show_manufacturing_process()

    # -----------------------------
    # 旧：計画計算＋表示（既存機能を踏襲）
    # -----------------------------
    def _show_plan_simulation(self):
        st.subheader("📊 計画シミュレーション")
        st.write("指定した期間の生産計画を計算・表示します。")

        col1, col2, col3 = st.columns([2, 2, 1])
        with col1:
            start_date = st.date_input(
                "開始日", datetime.now().date(),
                help="計画の開始日を選択してください"
            )
        with col2:
            end_date = st.date_input(
                "終了日", datetime.now().date() + timedelta(days=30),
                help="計画の終了日を選択してください"
            )
        with col3:
            st.write(""); st.write("")
            calculate_clicked = st.button("🔧 計画計算", type="primary", use_container_width=True)

        if calculate_clicked:
            self._calculate_and_show_plan(start_date, end_date)

    def _calculate_and_show_plan(self, start_date, end_date):
        with st.spinner("生産計画を計算中..."):
            try:
                plans = self.service.calculate_production_plan(start_date, end_date)
                if plans:
                    # DataFrame 化
                    plan_df = pd.DataFrame([{
                        'date': plan.date,
                        'product_id': plan.product_id,
                        'product_code': plan.product_code,
                        'product_name': plan.product_name,
                        'demand_quantity': plan.demand_quantity,
                        'planned_quantity': plan.planned_quantity,
                        'inspection_category': plan.inspection_category,
                        'is_constrained': plan.is_constrained
                    } for plan in plans])

                    self._display_production_plan(plan_df)
                else:
                    st.warning("指定期間内に生産計画データがありません")

            except Exception as e:
                st.error(f"計画計算エラー: {e}")

    def _display_production_plan(self, plan_df: pd.DataFrame):
        # サマリー
        st.subheader("📈 計画サマリー")
        total_demand = plan_df['demand_quantity'].sum()
        total_planned = plan_df['planned_quantity'].sum()
        constrained_count = plan_df['is_constrained'].sum()

        col1, col2, col3, col4 = st.columns(4)
        with col1: st.metric("総需要量", f"{total_demand:,.0f}")
        with col2: st.metric("総計画生産量", f"{total_planned:,.0f}")
        with col3:
            utilization = (total_planned / total_demand * 100) if total_demand > 0 else 0
            st.metric("計画達成率", f"{utilization:.1f}%")
        with col4: st.metric("制約対象製品数", int(constrained_count))

        # グラフ
        st.subheader("📊 生産計画チャート")
        fig = self.charts.create_production_plan_chart(plan_df)
        if fig:
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("グラフデータがありません")

        # 日次サマリー
        st.subheader("📋 詳細生産計画")
        st.write("**日次計画サマリー**")
        daily_summary = plan_df.groupby('date').agg({
            'demand_quantity': 'sum',
            'planned_quantity': 'sum'
        }).reset_index()
        daily_summary['達成率'] = (daily_summary['planned_quantity'] / daily_summary['demand_quantity'] * 100).round(1)

        st.dataframe(
            daily_summary,
            column_config={
                "date": "日付",
                "demand_quantity": st.column_config.NumberColumn("需要量", format="%d"),
                "planned_quantity": st.column_config.NumberColumn("計画生産量", format="%d"),
                "達成率": st.column_config.NumberColumn("達成率", format="%.1f%%"),
            },
            use_container_width=True,
        )

        # 製品別詳細
        st.write("**製品別詳細計画**")
        st.dataframe(
            plan_df,
            column_config={
                "date": "日付",
                "product_code": "製品コード",
                "product_name": "製品名",
                "demand_quantity": st.column_config.NumberColumn("需要量", format="%d"),
                "planned_quantity": st.column_config.NumberColumn("計画生産量", format="%d"),
                "inspection_category": "検査区分",
                "is_constrained": st.column_config.CheckboxColumn("制約対象"),
            },
            use_container_width=True,
        )

        # CSV 出力
        st.subheader("💾 データ出力")
        csv = plan_df.to_csv(index=False, encoding="utf-8-sig")
        st.download_button(
            label="📥 生産計画をCSVダウンロード",
            data=csv,
            file_name=f"production_plan_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            type="primary",
        )

    # -----------------------------
    # 新規：CRUD 管理タブ
    # -----------------------------
    def _show_plan_management(self):
        st.subheader("📝 生産計画管理（登録・更新・削除）")

        # --- 新規登録フォーム（最低限の項目をインライン実装） ---
        with st.form("create_production_form"):
            st.write("新しい計画を登録")
            product_id = st.number_input("製品ID", min_value=1, step=1)
            quantity = st.number_input("数量", min_value=1, step=1)
            scheduled_date = st.date_input("日付", value=date.today())
            submitted = st.form_submit_button("登録")

            if submitted:
                if hasattr(self.service, "create_production"):
                    payload = {
                        "product_id": int(product_id),
                        "quantity": int(quantity),
                        "scheduled_date": scheduled_date,
                    }
                    ok = self.service.create_production(payload)
                    if ok:
                        st.success("生産計画を登録しました")
                        st.rerun()
                    else:
                        st.error("生産計画の登録に失敗しました")
                else:
                    st.warning("create_production() が service に未実装です")

        # --- 一覧＆編集／削除 ---
        st.subheader("登録済み計画一覧")
        if not hasattr(self.service, "get_productions"):
            st.info("get_productions() が service に未実装です")
            return

        plans = self.service.get_productions()
        if not plans:
            st.info("登録されている生産計画はありません")
            return

        for plan in plans:
            with st.expander(f"📝 計画ID: {plan.id}"):
                st.write(f"製品ID: {plan.product_id}, 数量: {plan.quantity}, 日付: {plan.scheduled_date}")

                # 編集フォーム
                with st.form(f"edit_production_{plan.id}"):
                    new_product_id = st.number_input("製品ID", min_value=1, value=plan.product_id, key=f"p_{plan.id}")
                    new_quantity   = st.number_input("数量",    min_value=1, value=plan.quantity,    key=f"q_{plan.id}")
                    new_date       = st.date_input("日付", value=plan.scheduled_date, key=f"d_{plan.id}")

                    update_clicked = st.form_submit_button("更新")
                    if update_clicked:
                        if hasattr(self.service, "update_production"):
                            update_data = {
                                "product_id": int(new_product_id),
                                "quantity": int(new_quantity),
                                "scheduled_date": new_date,
                            }
                            ok = self.service.update_production(plan.id, update_data)
                            if ok:
                                st.success("計画を更新しました")
                                st.rerun()
                            else:
                                st.error("計画更新に失敗しました")
                        else:
                            st.warning("update_production() が service に未実装です")

                # 削除ボタン
                delete_clicked = st.button("🗑️ 削除", key=f"del_{plan.id}")
                if delete_clicked:
                    if hasattr(self.service, "delete_production"):
                        ok = self.service.delete_production(plan.id)
                        if ok:
                            st.success("計画を削除しました")
                            st.rerun()
                        else:
                            st.error("計画削除に失敗しました")
                    else:
                        st.warning("delete_production() が service に未実装です")

    # -----------------------------
    # 製造工程タブ
    # -----------------------------
    def _show_manufacturing_process(self):
        """製造工程（加工対象）- 積載計画数を基に表示"""
        st.subheader("🔧 製造工程（加工対象）")
        st.write("積載計画で設定された数量を製造工程の加工対象として表示します。")

        # 日付範囲選択
        st.markdown("---")
        st.subheader("📅 出力期間設定")
        col1, col2 = st.columns(2)

        with col1:
            start_date = st.date_input(
                "開始日",
                value=date.today(),
                key="mfg_start_date"
            )

        with col2:
            end_date = st.date_input(
                "終了日",
                value=date.today() + timedelta(days=7),
                key="mfg_end_date"
            )

        if start_date > end_date:
            st.error("開始日は終了日より前の日付を指定してください")
            return

        # データ取得
        try:
            # transport_serviceが必要
            if self.transport_service is None:
                st.error("transport_serviceが設定されていません。")
                return

            progress_df = self.transport_service.delivery_progress_repo.get_delivery_progress(start_date, end_date)

            if progress_df.empty:
                st.info("指定期間内にデータがありません")
                return

            # planned_quantityが設定されているデータのみ抽出
            if 'planned_quantity' not in progress_df.columns:
                st.warning("planned_quantity列がデータに含まれていません")
                return

            # planned_quantityが0より大きいものだけ表示
            progress_df = progress_df[progress_df['planned_quantity'] > 0].copy()

            if progress_df.empty:
                st.warning("指定期間内に計画数量が設定されているデータがありません")
                st.info("💡 配送便計画画面で積載計画を作成し、DBに保存してください")
                return

            # 日付を正規化
            progress_df['delivery_date'] = pd.to_datetime(progress_df['delivery_date']).dt.date

            # マトリックス表示
            st.markdown("---")
            st.subheader("📊 製品コード × 日付 マトリックス（加工対象数量）")

            matrix_df = self._create_matrix_view(progress_df)

            # データ表示
            st.dataframe(
                matrix_df,
                use_container_width=True,
                hide_index=False,
                height=600
            )

            # Excel出力
            st.markdown("---")
            st.subheader("💾 Excel出力")

            if st.button("📥 Excelダウンロード", type="primary", key="mfg_excel_download"):
                excel_data = self._export_manufacturing_to_excel(matrix_df, start_date, end_date)

                filename = f"製造工程_加工対象_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"

                st.download_button(
                    label="⬇️ ダウンロード開始",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="mfg_excel_download_btn"
                )

        except Exception as e:
            st.error(f"データ取得エラー: {e}")
            import traceback
            st.code(traceback.format_exc())

    def _create_matrix_view(self, progress_df: pd.DataFrame) -> pd.DataFrame:
        """製品コード×日付のマトリックスを作成（縦軸=製品コード、横軸=日付）"""

        # 製品コード一覧を取得（ソート）
        product_codes = sorted(progress_df['product_code'].unique())

        # 日付一覧を取得（ソート）
        dates = sorted(progress_df['delivery_date'].unique())
        date_columns = [d.strftime('%Y-%m-%d') for d in dates]

        # ピボットテーブル作成
        matrix_data = []

        for product_code in product_codes:
            product_data = progress_df[progress_df['product_code'] == product_code]

            # 製品名を取得
            product_name = product_data['product_name'].iloc[0] if not product_data.empty else ''

            row = {
                '製品コード': product_code,
                '製品名': product_name
            }

            # 各日付の計画数量を設定
            for date_obj, date_str in zip(dates, date_columns):
                day_data = product_data[product_data['delivery_date'] == date_obj]

                if not day_data.empty:
                    # 同じ製品コード・日付で複数レコードがある場合は合計
                    planned_qty = day_data['planned_quantity'].sum()
                    row[date_str] = int(planned_qty) if planned_qty > 0 else 0
                else:
                    row[date_str] = 0

            matrix_data.append(row)

        # DataFrameに変換
        matrix_df = pd.DataFrame(matrix_data)

        # インデックスを製品コードに設定
        matrix_df = matrix_df.set_index('製品コード')

        return matrix_df

    def _export_manufacturing_to_excel(self, matrix_df: pd.DataFrame, start_date: date, end_date: date) -> BytesIO:
        """マトリックスデータをExcelに出力"""

        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # データをExcelに書き込み
            matrix_df.to_excel(writer, sheet_name='加工対象', index=True)

            # ワークブックとシートを取得
            workbook = writer.book
            worksheet = writer.sheets['加工対象']

            # スタイル設定
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            center_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # ヘッダー行のスタイル設定（1行目）
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            # データ行のスタイル設定
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                          min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = border

                    # 数値セルの書式設定
                    if isinstance(cell.value, (int, float)) and cell.column > 2:
                        cell.number_format = '#,##0'

            # 列幅の自動調整
            worksheet.column_dimensions['A'].width = 15  # 製品コード
            worksheet.column_dimensions['B'].width = 30  # 製品名

            # 日付列の幅を設定
            for col_idx in range(3, worksheet.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 12

            # タイトル行を挿入
            worksheet.insert_rows(1)
            worksheet['A1'] = f"製造工程 加工対象一覧（{start_date.strftime('%Y年%m月%d日')} ～ {end_date.strftime('%Y年%m月%d日')}）"
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=worksheet.max_column)

        output.seek(0)
        return output
