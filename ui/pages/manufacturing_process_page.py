# app/ui/pages/manufacturing_process_page.py
import streamlit as st
import pandas as pd
from datetime import date, timedelta
from typing import Dict, Any
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ManufacturingProcessPage:
    """製造工程画面 - 積載計画数を基に加工対象を表示"""

    def __init__(self, transport_service):
        self.service = transport_service

    def show(self):
        """ページ表示"""
        st.title("🏭 製造工程（加工対象）")
        st.write("積載計画で設定された数量を製造工程の加工対象として表示します。")

        # 日付範囲選択
        st.subheader("📅 出力期間設定")
        col1, col2 = st.columns(2)

        with col1:
            start_date = st.date_input(
                "開始日",
                value=date.today(),
                key="mfg_start_date"
            )

        with col2:
            end_date = st.date_input(
                "終了日",
                value=date.today() + timedelta(days=7),
                key="mfg_end_date"
            )

        if start_date > end_date:
            st.error("開始日は終了日より前の日付を指定してください")
            return

        # データ取得
        try:
            progress_df = self.service.get_delivery_progress(start_date, end_date)

            if progress_df.empty:
                st.info("指定期間内にデータがありません")
                return

            # planned_quantityが設定されているデータのみ抽出
            if 'planned_quantity' not in progress_df.columns:
                st.warning("planned_quantity列がデータに含まれていません")
                return

            # planned_quantityが0より大きいものだけ表示
            progress_df = progress_df[progress_df['planned_quantity'] > 0].copy()

            if progress_df.empty:
                st.warning("指定期間内に計画数量が設定されているデータがありません")
                st.info("💡 配送便計画画面で積載計画を作成し、DBに保存してください")
                return

            # 日付を正規化
            progress_df['delivery_date'] = pd.to_datetime(progress_df['delivery_date']).dt.date

            # マトリックス表示
            st.subheader("📊 製品コード × 日付 マトリックス（加工対象数量）")

            matrix_df = self._create_matrix_view(progress_df)

            # データ表示
            st.dataframe(
                matrix_df,
                use_container_width=True,
                hide_index=False,
                height=600
            )

            # Excel出力
            st.markdown("---")
            st.subheader("💾 Excel出力")

            if st.button("📥 Excelダウンロード", type="primary"):
                excel_data = self._export_to_excel(matrix_df, start_date, end_date)

                filename = f"製造工程_加工対象_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"

                st.download_button(
                    label="⬇️ ダウンロード開始",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"データ取得エラー: {e}")
            import traceback
            st.code(traceback.format_exc())

    def _create_matrix_view(self, progress_df: pd.DataFrame) -> pd.DataFrame:
        """製品コード×日付のマトリックスを作成（縦軸=製品コード、横軸=日付）"""

        # 製品コード一覧を取得（ソート）
        product_codes = sorted(progress_df['product_code'].unique())

        # 日付一覧を取得（ソート）
        dates = sorted(progress_df['delivery_date'].unique())
        date_columns = [d.strftime('%Y-%m-%d') for d in dates]

        # ピボットテーブル作成
        matrix_data = []

        for product_code in product_codes:
            product_data = progress_df[progress_df['product_code'] == product_code]

            # 製品名を取得
            product_name = product_data['product_name'].iloc[0] if not product_data.empty else ''

            row = {
                '製品コード': product_code,
                '製品名': product_name
            }

            # 各日付の計画数量を設定
            for date_obj, date_str in zip(dates, date_columns):
                day_data = product_data[product_data['delivery_date'] == date_obj]

                if not day_data.empty:
                    # 同じ製品コード・日付で複数レコードがある場合は合計
                    planned_qty = day_data['planned_quantity'].sum()
                    row[date_str] = int(planned_qty) if planned_qty > 0 else 0
                else:
                    row[date_str] = 0

            matrix_data.append(row)

        # DataFrameに変換
        matrix_df = pd.DataFrame(matrix_data)

        # インデックスを製品コードに設定
        matrix_df = matrix_df.set_index('製品コード')

        return matrix_df

    def _export_to_excel(self, matrix_df: pd.DataFrame, start_date: date, end_date: date) -> BytesIO:
        """マトリックスデータをExcelに出力"""

        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # データをExcelに書き込み
            matrix_df.to_excel(writer, sheet_name='加工対象', index=True)

            # ワークブックとシートを取得
            workbook = writer.book
            worksheet = writer.sheets['加工対象']

            # スタイル設定
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            center_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # ヘッダー行のスタイル設定（1行目）
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            # データ行のスタイル設定
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                          min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = border

                    # 数値セルの書式設定
                    if isinstance(cell.value, (int, float)) and cell.column > 2:
                        cell.number_format = '#,##0'

            # 列幅の自動調整
            worksheet.column_dimensions['A'].width = 15  # 製品コード
            worksheet.column_dimensions['B'].width = 30  # 製品名

            # 日付列の幅を設定
            for col_idx in range(3, worksheet.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 12

            # タイトル行を挿入
            worksheet.insert_rows(1)
            worksheet['A1'] = f"製造工程 加工対象一覧（{start_date.strftime('%Y年%m月%d日')} ～ {end_date.strftime('%Y年%m月%d日')}）"
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=worksheet.max_column)

        output.seek(0)
        return output
